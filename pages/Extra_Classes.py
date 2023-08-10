import os
import time
import streamlit as st
from dotenv import load_dotenv
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
from io import BytesIO
from openpyxl.styles import Font
from bs4 import BeautifulSoup

# getting credentials from environment variables
load_dotenv()
MANAGER_USERNAME = os.getenv("MANAGER_USERNAME")
MANAGER_PASSWORD = os.getenv("MANAGER_PASSWORD")
CHROMEDRIVER_PATH = os.environ.get("CHROMEDRIVER_PATH")
GOOGLE_CHROME_BIN = os.environ.get("GOOGLE_CHROME_BIN")

if 'data' not in st.session_state:
    st.session_state.data = []
    st.session_state.option = []

st.set_page_config(
    page_title="Tri Viet Education",
    page_icon=":pager:",
    initial_sidebar_state="expanded",
)

def write_to_excel(class_name, students):
    output = BytesIO()
    wb_obj = openpyxl.load_workbook("TUTORING COURSE OUTLINE.xlsx")
    sheet_obj = wb_obj.active 
    # modify title into appropriate class
    name = sheet_obj.cell(row = 1, column = 1) 
    name.value = f'{class_name} - EXTRA CLASS'
    # modify student columns
    # col pointer starts at 3 (hard coded atm)
    col_pointer = 4
    # setting student detail font to bold
    student_font = Font(name="Time News Roman", size=12, bold=True)
    for i, student in enumerate(students):
        # writing student index
        index_cell = sheet_obj.cell(2,col_pointer)
        index_cell.value = i + 1
        index_cell.font = student_font
        # writing students' names
        if type(sheet_obj.cell(3, col_pointer)).__name__ == 'MergedCell':
            student_cell = sheet_obj.unmerge_cells(start_row=3,start_column=col_pointer, end_row=3, end_column=col_pointer+2)
        student_cell = sheet_obj.cell(3,col_pointer)
        student_cell.value = student
        student_cell.font = student_font
        student_cell = sheet_obj.merge_cells(start_row=3,start_column=col_pointer, end_row=3, end_column=col_pointer+2)
        # writing skill cells
        listening_cell = sheet_obj.cell(4,col_pointer)
        listening_cell.value = "List."
        listening_cell.font = student_font
        reading_writing_cell = sheet_obj.cell(4,col_pointer+1)
        reading_writing_cell.value = "R&W"
        reading_writing_cell.font = student_font
        vocab_cell = sheet_obj.cell(4,col_pointer+2)
        vocab_cell.value = "Vocab"
        vocab_cell.font = student_font
        col_pointer = col_pointer+3
    wb_obj.save(output)
    return output.getvalue()

@st.cache_resource
def load_data():
    # initialize the Chrome driver
    options = Options()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.binary_location = GOOGLE_CHROME_BIN
    options.add_argument('--headless')
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    driver = webdriver.Chrome(options=options, service=ChromeService(ChromeDriverManager(version="114.0.5735.16").install()))
    # login page
    driver.get("https://trivietedu.ileader.vn/login.aspx")
    # find username/email field and send the username itself to the input field
    driver.find_element("id","user").send_keys(MANAGER_USERNAME)
    # find password input field and insert password as well
    driver.find_element("id","pass").send_keys(MANAGER_PASSWORD)
    # click login button
    driver.find_element(By.XPATH,'//*[@id="login"]/button').click()
    driver.get("https://trivietedu.ileader.vn/Default.aspx?mod=lophoc!lophoc")
    time.sleep(15)
    soup = BeautifulSoup(driver.page_source,"lxml")
    classes = {}
    class_titles = soup.find('tbody').find_all('tr')
    st.write(class_titles)
    for class_title in class_titles:
        classes[[c.text for c in class_title][1]] = [c['href'] for c in class_title.find_all(title="Xem danh sách lớp học")][0]
    st.write(classes)
    return driver, classes, soup

if st.button("Refresh"):
    st.cache_resource.clear()

driver, classes, soup = load_data()

options = st.multiselect(
    'Select your classes:',
    classes.keys(),
    list(classes.keys())[-1])


placeholder = st.empty()

if options:
    confirm = placeholder.button('Select', key = 1)
else:
    confirm = placeholder.button('Select', disabled=True, key = 2)

data = {}

if confirm:
    st.session_state.data = []
    st.session_state.option = []
    for option in options:
        driver.execute_script(classes[option])
        time.sleep(2)
        student_soup = BeautifulSoup(driver.page_source, "lxml")
        table = [soup.find_all('tr') for soup in student_soup.find_all("tbody")][1]
        rows= [row.find_all('td') for row in table]
        students = [student[2].text for student in rows]
        st.session_state.data.append(write_to_excel(option, students))
        st.session_state.option.append(option)
    placeholder.empty()

for i in range(len(st.session_state.data)):
    st.download_button(label=f'📥 Download {st.session_state.option[i]}',
                        data=st.session_state.data[i],
                        file_name= f'{st.session_state.option[i]}-EXTRA CLASS.xlsx')

