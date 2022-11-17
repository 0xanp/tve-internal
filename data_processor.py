import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import pandas as pd

# check if template folder is empty else delete all old templates
template_path = r'\\pdt-ngocle\Training Data\EXTRA CLASSES\TEMPLATES'
template_list = os.listdir(template_path) 
if template_list:
    for template in template_list:
        os.remove(f'{template_path}\{template}')

# getting all file names from student list directory 
students_path = "./student-lists"
dir_list = os.listdir(students_path) 

for list in dir_list:
    # get student data as a data frame
    student_data = pd.read_html(f'{students_path}/{list}')[0]
    # reading from templates
    template_path = './templates/TUTORING COURSE OUTLINE - KET 1.xlsx'
    output_path = r'\\pdt-ngocle\Training Data\EXTRA CLASSES\TEMPLATES'
    wb_obj = openpyxl.load_workbook(template_path)
    # generate new file names based on student lists
    file_name = f'{output_path}\{list.split(".")[0]} - EXTRA CLASS.xlsx'
    # get active sheet from template
    sheet_obj = wb_obj.active 
    # modify title into appropriate class
    name = sheet_obj.cell(row = 1, column = 1) 
    name.value = f'EXTRA CLASS {list.split(".")[0]}'
    # modify student columns
    # col pointer starts at 3 (hard coded atm)
    col_pointer = 4
    # setting student detail font to bold
    student_font = Font(name="Time News Roman", size=12, bold=True)
    for i, student in enumerate(student_data['Tên']):
        # writing student index
        index_cell = sheet_obj.cell(2,col_pointer)
        index_cell.value = i + 1
        index_cell.font = student_font
        # writing students' names
        if type(sheet_obj.cell(3, col_pointer)).__name__ == 'MergedCell':
            student_cell = sheet_obj.unmerge_cells(start_row=3,start_column=col_pointer, end_row=3, end_column=col_pointer+2)
        student_cell = sheet_obj.cell(3,col_pointer)
        student_cell.value = student
        student_cell.font = student_font
        student_cell = sheet_obj.merge_cells(start_row=3,start_column=col_pointer, end_row=3, end_column=col_pointer+2)
        # writing skill cells
        listening_cell = sheet_obj.cell(4,col_pointer)
        listening_cell.value = "List."
        listening_cell.font = student_font
        reading_writing_cell = sheet_obj.cell(4,col_pointer+1)
        reading_writing_cell.value = "R&W"
        reading_writing_cell.font = student_font
        vocab_cell = sheet_obj.cell(4,col_pointer+2)
        vocab_cell.value = "Vocab"
        vocab_cell.font = student_font
        col_pointer = col_pointer+3
    wb_obj.save(file_name)
